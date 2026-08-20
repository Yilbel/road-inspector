"""
6_final_report.py
------------------
5_dedupe_and_prioritize.py'nin ürettiği tekilleştirilmiş, öncelik
puanlı listeden BELEDİYEYE SUNULACAK final raporu üretir:
  - Her benzersiz nesne için tek bir işaretli görsel (temsili kare)
  - Öncelik sütunu içeren, kategoriye göre gruplu, profesyonel formatlı Excel raporu

Kullanım:
    python 6_final_report.py --frames output/frames --detections output/detections/tespitler_tekil.csv --out output/report
"""
import os
import ast
import argparse
import cv2
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_bbox(value):
    if isinstance(value, str):
        return ast.literal_eval(value)
    return value


PRIORITY_COLOR = {
    "yüksek": (0, 0, 255),   # kırmızı (BGR - OpenCV çizim için)
    "orta": (0, 140, 255),   # turuncu
    "düşük": (0, 255, 255),  # sarı
    "bilgi": (255, 0, 0),    # mavi
}

# Excel hücre renklendirmesi için (ARGB hex, openpyxl formatı)
PRIORITY_FILL_HEX = {
    "yüksek": "FFC7CE",  # açık kırmızı
    "orta": "FFE5B4",    # açık turuncu
    "düşük": "FFF2CC",   # açık sarı
    "bilgi": "DDEBF7",   # açık mavi
}
PRIORITY_FONT_HEX = {
    "yüksek": "9C0006",
    "orta": "9C5700",
    "düşük": "9C8A00",
    "bilgi": "1F4E78",
}

HEADER_FILL = "2F5597"   # koyu mavi
HEADER_FONT = "FFFFFF"   # beyaz


def draw_and_save(frames_dir: str, df: pd.DataFrame, out_dir: str):
    annotated_dir = os.path.join(out_dir, "isaretli_kareler")
    os.makedirs(annotated_dir, exist_ok=True)
    for idx, row in df.iterrows():
        frame_name = row["temsili_kare"]
        frame_path = os.path.join(frames_dir, frame_name)
        img = cv2.imread(frame_path)
        if img is None:
            continue
        bbox = parse_bbox(row["bbox_x1y1x2y2"])
        x1, y1, x2, y2 = map(int, bbox)
        color = PRIORITY_COLOR.get(row["oncelik"], (0, 255, 0))
        cv2.rectangle(img, (x1, y1), (x2, y2), color, 3)
        label = f"{row['kategori']} | {row['oncelik']} | {row['gorulme_sayisi']}x gorulmus"
        cv2.putText(img, label, (x1, max(y1 - 12, 0)), cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)
        # Dosya adını nesne no'suyla benzersiz yapalım (aynı kare birden fazla nesne için temsili olabilir)
        out_name = f"nesne_{idx+1:03d}_{row['kategori'].replace(' ', '_').replace('(', '').replace(')', '')}.jpg"
        cv2.imwrite(os.path.join(annotated_dir, out_name), img)
    print(f"İşaretli kareler -> {annotated_dir}")


def _style_header(ws, n_cols: int):
    """Başlık satırını kalın/renkli yapar ve üstte sabitler (freeze panes)."""
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    header_font = Font(bold=True, color=HEADER_FONT, size=11)
    thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def _autofit_columns(ws, df: pd.DataFrame, max_width: int = 45):
    """Sütun genişliklerini içeriğe göre otomatik ayarlar."""
    for i, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(i)
        header_len = len(str(col))
        max_data_len = df[col].astype(str).map(len).max() if len(df) > 0 else 0
        width = min(max(header_len, max_data_len) + 3, max_width)
        ws.column_dimensions[col_letter].width = width


def _color_priority_column(ws, df: pd.DataFrame, priority_col_name: str = "oncelik"):
    """'oncelik' sütunundaki hücreleri değerine göre renklendirir."""
    if priority_col_name not in df.columns:
        return
    col_idx = list(df.columns).index(priority_col_name) + 1
    col_letter = get_column_letter(col_idx)
    for row_i, value in enumerate(df[priority_col_name], start=2):
        fill_hex = PRIORITY_FILL_HEX.get(value)
        font_hex = PRIORITY_FONT_HEX.get(value)
        if fill_hex:
            cell = ws[f"{col_letter}{row_i}"]
            cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            cell.font = Font(bold=True, color=font_hex)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_zebra_and_alignment(ws, df: pd.DataFrame):
    """Okunabilirlik için satırları hafif zebra desenli yapar, tüm hücreleri ortalar (metin sola)."""
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    n_cols = len(df.columns)
    for row_i in range(2, len(df) + 2):
        if row_i % 2 == 0:
            for col_i in range(1, n_cols + 1):
                cell = ws.cell(row=row_i, column=col_i)
                if cell.fill.start_color.rgb in (None, "00000000"):
                    cell.fill = zebra_fill
        for col_i in range(1, n_cols + 1):
            cell = ws.cell(row=row_i, column=col_i)
            cell.alignment = Alignment(vertical="center")


def _finalize_sheet(ws, df: pd.DataFrame, priority_col_name: str = None):
    n_cols = len(df.columns)
    _style_header(ws, n_cols)
    _autofit_columns(ws, df)
    _apply_zebra_and_alignment(ws, df)
    if priority_col_name:
        _color_priority_column(ws, df, priority_col_name=priority_col_name)


def build_excel(df: pd.DataFrame, out_dir: str):
    excel_path = os.path.join(out_dir, "belediye_raporu_final.xlsx")

    summary = df.groupby("kategori").size().reset_index(name="adet").sort_values("adet", ascending=False)
    summary.columns = ["Kategori", "Adet"]

    road_damage = df[df["kategori"].str.startswith("yol hasarı")].copy()
    road_damage_priority = (
        road_damage.groupby("oncelik").size().reset_index(name="adet")
        if len(road_damage) > 0 else pd.DataFrame(columns=["oncelik", "adet"])
    )
    road_damage_priority.columns = ["Öncelik", "Adet"]

    display_cols = [
        "kategori", "oncelik", "gorulme_sayisi", "en_yuksek_guven",
        "ilk_gorulme_sn", "son_gorulme_sn", "temsili_kare"
    ]
    has_gps = "enlem" in df.columns and "boylam" in df.columns
    if has_gps:
        display_cols += ["enlem", "boylam", "google_maps"]
    display_df = df[display_cols].sort_values(["kategori", "oncelik"]).reset_index(drop=True)
    display_df.columns = [
        "Kategori", "Öncelik", "Görülme Sayısı", "En Yüksek Güven",
        "İlk Görülme (sn)", "Son Görülme (sn)", "Temsili Kare"
    ] + (["Enlem", "Boylam", "Google Maps"] if has_gps else [])

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Özet", index=False)
        road_damage_priority.to_excel(writer, sheet_name="Yol Hasarı Önceliği", index=False)
        display_df.to_excel(writer, sheet_name="Tüm Nesneler", index=False)

        wb = writer.book
        _finalize_sheet(wb["Özet"], summary)
        _finalize_sheet(wb["Yol Hasarı Önceliği"], road_damage_priority, priority_col_name="Öncelik")
        _finalize_sheet(wb["Tüm Nesneler"], display_df, priority_col_name="Öncelik")

    print(f"Excel raporu -> {excel_path}" + (" (konum bilgisiyle)" if has_gps else ""))
    return summary


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--frames", default="output/frames")
    parser.add_argument("--detections", default="output/detections/tespitler_gps.csv",
                         help="GPS eklenmiş dosya varsa onu, yoksa tespitler_tekil.csv'yi kullan")
    parser.add_argument("--out", default="output/report")
    args = parser.parse_args()

    os.makedirs(args.out, exist_ok=True)
    df = pd.read_csv(args.detections)

    summary = build_excel(df, args.out)
    draw_and_save(args.frames, df, args.out)

    print(f"\nToplam {len(df)} benzersiz nesne.")
    print(summary.to_string(index=False))